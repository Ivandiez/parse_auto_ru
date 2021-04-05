import requests
import os
import xlsxwriter
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import pandas as pd


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    writer.close()


def read_file(filename):
    with open(filename) as input_file:
        text = input_file.read()
    return text


def parse_data_bs(filename):
    results = []
    text = read_file(filename)

    soup = BeautifulSoup(text, features="lxml")

    car_desc = soup.find_all('div', {'class': 'ListingItem-module__description'})
    for item in car_desc:
        car_name = \
            item.find('h3', {'class': 'ListingItemTitle-module__container ListingItem-module__title'}
                      ).find('a').text
        car_link = \
            item.find('h3', {'class': 'ListingItemTitle-module__container ListingItem-module__title'}
                      ).find('a').get('href')
        car_tech_summary = \
            item.find_all('div', {'class': 'ListingItemTechSummaryDesktop__cell'})

        # строка - 1.8 л. / 125 л.с. / Бензин
        car_engine_volume = car_tech_summary[0].text.split('/')[0]  # 1.8 л.
        car_engine_horse_volume = car_tech_summary[0].text.split('/')[1]  # 125 л.с.
        car_engine_type = car_tech_summary[0].text.split('/' )[2]  # Бензин

        car_transmition_type = car_tech_summary[1].text # Механика
        car_type = car_tech_summary[2].text  # седан, хэтчбек, универсал, купе, внедорожник,
        # минивэн, пикап, кабриолет, фургон
        car_wheel_drive = car_tech_summary[3].text  # передний, задний, полный
        car_color = car_tech_summary[4].text  # цвет

        if item.find('div', {'class': 'ListingItemPrice-module__content'}):
            car_price = item.find('div', {'class': 'ListingItemPrice-module__content'}).text[:-2]
        else:
            car_price = None
        car_year = item.find('div', {'class': 'ListingItem-module__year'}).text
        car_miliage = item.find('div', {'class': 'ListingItem-module__kmAge'}).text[:-3]

        results.append({
            'car_name': car_name,
            'car_link': car_link,
            'car_engine_volume': car_engine_volume,
            'car_engine_horse_volume': car_engine_horse_volume,
            'car_engine_type': car_engine_type,
            'car_transmition_type': car_transmition_type,
            'car_type': car_type,
            'car_wheel_drive': car_wheel_drive,
            'car_color': car_color,
            'car_price, Р': car_price,
            'car_year': car_year,
            'car_miliage, км': car_miliage
    })
    return results


for i in range(94):
    with open('test.html', 'w', encoding='utf-8') as output_file:
        url = "https://auto.ru/moskva/cars/vendor-foreign/all/do-300000/" \
              f"?page={i + 1}&sort=cr_date-desc&transmission=MECHANICAL&year_from=2005"

        r = requests.get(url)
        r.encoding = 'utf-8'
        output_file.write(r.text)
    car_s = parse_data_bs('test.html')
    car_df = pd.DataFrame(car_s)
    if i != 0:
        append_df_to_excel('/home/ivan/PycharmProjects/parse_auto_from_auto_ru/output.xlsx',
                           car_df, header=None, index=None)
    else:
        append_df_to_excel('/home/ivan/PycharmProjects/parse_auto_from_auto_ru/output.xlsx',
                           car_df, header=None, index=None)
    #car_df.to_excel('output.xlsx')
    print(car_df)

workbook = xlsxwriter.workbook('output.xlsx')
worksheet = workbook.worksheets()

for i in range(len())
worksheet.write()